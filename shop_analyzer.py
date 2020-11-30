from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, PatternFill, Font
import xlrd
from datetime import datetime
from tkinter import filedialog, messagebox
import tkinter as tk
from tkinter import *
import xlsxwriter
import sys

# cxfreeze-quickstart for creating exe, nsis for creating installer

COLUMN_SETNO_NAME = 4 - 1
COLUMN_QTY = 11 - 1
COLUMN_INCOME = 13 - 1

RESULT_PROD_NAME = 0
RESULT_SELLS_QTY = 1
RESULT_PRICE = 2
RESULT_SUM = 3

OFFSET = 3
file_paths = " "
first_day = " "
last_day = " "

product_info = {
    'dolkifrukt100g': {'pos': 0, 'qty': 0, 'price': 0, 'income': 0},
    'maramelkiklub60g': {'pos': 1, 'qty': 0, 'price': 0, 'income': 0},
    'maramelki_malina60g': {'pos': 2, 'qty': 0, 'price': 0, 'income': 0},
    'maramelkiyabl-baz60g': {'pos': 3, 'qty': 0, 'price': 0, 'income': 0},
    'maramelki_grusha-imbyr': {'pos': 4, 'qty': 0, 'price': 0, 'income': 0},
    'pastila_klub_50': {'pos': 5, 'qty': 0, 'price': 0, 'income': 0},
    'pastila_malina_50': {'pos': 6, 'qty': 0, 'price': 0, 'income': 0},
    'pechen_malina': {'pos': 7, 'qty': 0, 'price': 0, 'income': 0},
    'pechen_abrik': {'pos': 8, 'qty': 0, 'price': 0, 'income': 0},
    'pechen_sliva': {'pos': 9, 'qty': 0, 'price': 0, 'income': 0},
    'pechen_klub': {'pos': 10, 'qty': 0, 'price': 0, 'income': 0},
    'total': {'pos': 11, 'qty': 0, 'price': ' ', 'income': 0}
}

total_sells = product_info['total']

product_names = [
    'Дольки фруктовые',
    'Мармелад клубника',
    'Мармелад малина',
    'Мармелад яблоко-базилик',
    'Мармелад гурша-имбирь',
    'Пастила клубника',
    'Пастила малина',
    'Печенье малина',
    'Печенье абрикос',
    'Печенье слива',
    'Печенье клубника',
    'Общее'
]


def load_files():
    global file_paths
    file_paths = filedialog.askopenfilenames()
    files_loaded.insert(1.0, len(file_paths))


def update_value(prod_row):
    if prod_row[COLUMN_SETNO_NAME] in product_info:
        # + 1 to skip menu row
        product = product_info[prod_row[COLUMN_SETNO_NAME]]
        product['qty'] += prod_row[COLUMN_QTY]
        total_sells['qty'] += prod_row[COLUMN_QTY]
        product['income'] += prod_row[COLUMN_INCOME]
        total_sells['income'] += prod_row[COLUMN_INCOME]
        product['price'] = product['income'] / product['qty']

    else:
        print("Product code is not known. Error")


def handle_file(file_path):
    curr_file = xlrd.open_workbook(file_path, formatting_info=True)
    curr_data = curr_file.sheet_by_index(0)
    for rownum in range(curr_data.nrows):
        row = curr_data.row_values(rownum)
        # checking if it is product row (needed) or sum row/info row (should be skipped)
        if isinstance((row[COLUMN_INCOME]), float):
            if row[COLUMN_SETNO_NAME]:
                update_value(row)


def start_process():
    global file_paths, first_day, last_day
    first_day = day1.get()
    last_day = day2.get()
    for path in file_paths:
        handle_file(path)
    root.destroy()


root = tk.Tk()
root.geometry('450x350')

header_frame = Frame(root)
header_frame.pack()

desc_frame = Frame(root)
desc_frame.pack()

current_date_frame = Frame(root)
current_date_frame.pack()

find_file_btn_frame = Frame(root)
find_file_btn_frame.pack()

explanation_frame = Frame(root)
explanation_frame.pack()

calculate_btn_frame = Frame(root)
calculate_btn_frame.pack()

tk.Label(header_frame,
         text="Укажите временной интервал отчета \nЗатем выберите excel-отчеты (все сразу) и нажмите Рассчитать\nОтчет будет создан в папке запуска проекта с именем Отчет.xlsx",
         padx=5, pady=5).pack()

tk.Label(header_frame,
         text="Пример интервала: 2.10 - 15.10", padx=5, pady=5).pack()

tk.Label(current_date_frame,
         text="Интервал:").pack(side=tk.LEFT, padx=2, pady=5)

day1 = tk.Entry(current_date_frame, width=12, justify=CENTER)
day1.pack(side=tk.LEFT, pady=5)
day1.focus()

tk.Label(current_date_frame,
         text="-", width=2).pack(side=tk.LEFT, pady=2)
day2 = tk.Entry(current_date_frame, width=12, justify=CENTER)
day2.pack(side=tk.LEFT, pady=5)

find_first_file_btn = tk.Button(find_file_btn_frame,
                                text='Загрузить отчеты', command=load_files)
find_first_file_btn.pack(padx=5, pady=5)

tk.Label(explanation_frame,
         text="Загружено файлов: ", width=15).pack(side=tk.LEFT, pady=2)
files_loaded = Text(explanation_frame, height=1, width=5)
files_loaded.pack(side=tk.LEFT, pady=5)

calculate_btn = tk.Button(calculate_btn_frame,
                          text='Рассчитать', command=start_process)
calculate_btn.pack(padx=5, pady=5)

root.mainloop()

# result
result = {}
result_name = 'Всі свої. Звіт за ' + first_day + '-' + last_day + '.xlsx'
workbook = xlsxwriter.Workbook(result_name)
worksheet = workbook.add_worksheet('Data')
workbook.formats[0].set_font_size(14)
bold = workbook.add_format({'size': 14, 'bold': True, 'text_wrap': True, 'align': 'center', 'border': 1})
bordered = workbook.add_format({'size': 14, 'text_wrap': True, 'align': 'center', 'border': 1})
bold.set_align('vcenter')
colored = workbook.add_format({'size': 14, 'bold': True, 'bg_color': '#d9d9d9', 'border': True})
colored_n_right = workbook.add_format(
    {'size': 14, 'bold': True, 'bg_color': '#d9d9d9', 'border': True, 'align': 'right'})
right = workbook.add_format({'size': 14, 'align': 'right'})

worksheet.write(0, 2, 'Продажи с', bold)
worksheet.write(0, 3, first_day, bold)
worksheet.write(1, 2, 'По', bold)
worksheet.write(1, 3, last_day, bold)

row_num = 0 + OFFSET

worksheet.set_column(RESULT_PROD_NAME, RESULT_PROD_NAME, 25)
worksheet.set_column(RESULT_SELLS_QTY, RESULT_SELLS_QTY, 9)
worksheet.set_column(RESULT_PRICE, RESULT_PRICE, 9)
worksheet.set_column(RESULT_SUM, RESULT_SUM, 9)

worksheet.write(row_num, RESULT_PROD_NAME, 'Название', bold)
worksheet.write(row_num, RESULT_PROD_NAME, 'Название', bold)
worksheet.write(row_num, RESULT_SELLS_QTY, 'Кол-во продаж', bold)
worksheet.write(row_num, RESULT_PRICE, 'Цена', bold)
worksheet.write(row_num, RESULT_SUM, 'Стоимость, грн', bold)

for prod_name in product_names:
    row_num += 1
    worksheet.write(row_num, RESULT_PROD_NAME, prod_name, bold)


for key_name in product_info:
    prod = product_info[key_name]
    worksheet.write(prod['pos'] + OFFSET + 1, RESULT_SELLS_QTY, prod['qty'], bordered)
    worksheet.write(prod['pos'] + OFFSET + 1, RESULT_PRICE, prod['price'], bordered)
    worksheet.write(prod['pos'] + OFFSET + 1, RESULT_SUM, prod['income'], bordered)


#worksheet.autofilter(OFFSET, RESULT_PROD_NAME, row_num + 10, RESULT_SELLS_QTY)
#worksheet.freeze_panes(0. 1)
worksheet.add_table('A3:D11', {'header_row': False, 'autofilter': False, 'style': None})
workbook.close()



