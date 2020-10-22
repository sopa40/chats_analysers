from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from datetime import datetime
from tkinter import filedialog, messagebox
import tkinter as tk
from tkinter import *
import xlsxwriter
import sys

# cxfreeze-quickstart for creating exe, nsis for creating installer

old_file_path = ""
new_file_path = ""

new_day = " "


SEVEN_DAYS_IN_SECONDS = 568800
DAY_IN_SECONDS = 86400
TEN_HOURS_IN_SECONDS = 36000

days_past_in_sec = 0

def update_data():
    global new_day, days_past_in_sec
    day_two = int(day2.get())
    month_two = int(month2.get())
    year_two = int(year2.get())
    days_past_in_sec = (int(days_past.get()) * DAY_IN_SECONDS) - TEN_HOURS_IN_SECONDS
    if year_two < 2015 or day_two > 31 or month_two > 12 or day_two < 1 or month_two < 1:
        print("Wrong date input!")
        sys.exit()
    new_day = datetime(year_two, month_two, day_two)
    print(new_day)
    root.destroy()


def load_prev_file():
    global old_file_path
    old_file_path = filedialog.askopenfilename()


def load_current_file():
    global new_file_path
    new_file_path = filedialog.askopenfilename()


root = tk.Tk()
root.geometry('450x350')

header_frame = Frame(root)
header_frame.pack()

desc_frame = Frame(root)
desc_frame.pack()

current_date_frame = Frame(root)
current_date_frame.pack()

days_past_frame = Frame(root)
days_past_frame.pack()

explanation_frame = Frame(root)
explanation_frame.pack()

find_file_btn_frame = Frame(root)
find_file_btn_frame.pack()

calculate_btn_frame = Frame(root)
calculate_btn_frame.pack()

tk.Label(header_frame,
         text="Укажите дату текущего отчета \nЗатем выберите предыдущий и текущий excel-отчеты и нажмите Рассчитать\nОтчет будет создан в папке запуска проекта с именем Отчет.xlsx",
         padx=5, pady=5).pack()

tk.Label(header_frame,
         text="Пример даты: 2 4 2020", padx=5, pady=5).pack()

tk.Label(current_date_frame,
         text="Текущий отчет:").pack(side=tk.LEFT, pady=5)

tk.Label(current_date_frame,
         text="День", width=6).pack(side=tk.LEFT, pady=5)
day2 = tk.Entry(current_date_frame, width=3, justify=CENTER)
day2.pack(side=tk.LEFT, pady=5)
day2.focus()

tk.Label(current_date_frame,
         text="Месяц", width=6).pack(side=tk.LEFT, pady=5)
month2 = tk.Entry(current_date_frame, width=3, justify=CENTER)
month2.pack(side=tk.LEFT, pady=5)

tk.Label(current_date_frame,
         text="Год", width=4).pack(side=tk.LEFT, pady=5)
year2 = tk.Entry(current_date_frame, width=5, justify=CENTER)
year2.pack(side=tk.LEFT, pady=5)

tk.Label(days_past_frame,
         text="Дней с момента предыдущего отчета:", padx=5, pady=5).pack(side=tk.LEFT, pady=5)
days_past = tk.Entry(days_past_frame, width=5, justify=CENTER)
days_past.pack(side=tk.LEFT, pady=5)

tk.Label(explanation_frame,
         text="Например, с 15.10 по 20.10 прошло 5 дней.\n"
              " То есть текущий день не учитывать.", padx=5, pady=5).pack(pady=5)

find_first_file_btn = tk.Button(find_file_btn_frame,
                                text='Загрузить предыдущий отчет', command=load_prev_file)
find_first_file_btn.pack(padx=5, pady=5)
find_second_file_btn = tk.Button(find_file_btn_frame,
                                 text='Загрузить текущий отчет', command=load_current_file)
find_second_file_btn.pack(padx=5, pady=5)
calculate_btn = tk.Button(calculate_btn_frame,
                          text='Рассчитать', command=update_data)
calculate_btn.pack(padx=5, pady=5)

root.mainloop()


# TODO: Description, data input


new_file = load_workbook(new_file_path)
old_file = load_workbook(old_file_path)

new_data = new_file.worksheets[0]
old_data = old_file.worksheets[0]

COLUMN_SHOP = "B"
COLUMN_PRODUCT = "D"
COLUMN_LEFTOVER = "G"
COLUMN_DATE = "I"
COLUMN_INPUT = "J"

RESULT_SHOP_NAME = 0
RESULT_PROD_NAME = 1
RESULT_LEFTOVER = 2
RESULT_DATE = 3
RESULT_INPUT = 4
RESULT_SELLS = 5

new_shops = {}
old_shops = {}

old_row_num = 0
new_row_num = 0

if old_row_num != new_row_num:
    print("Rows in old and new are not equal. Exit")
    sys.exit()

product_names = {'Мармелад овощной', 'Дольки овощные', 'Паст50ЛТСетМалинСлив', 'Дольки фруктовые',
                 'Пастила М/С', 'Пастила К/А', 'Пастила Я/Г', 'Мармелад А/М/С', 'Мармелад клубника', 'Мармелад Я/Б'}


def add_item(sheet, shop, row):
    product_cell_name = "{}{}".format(COLUMN_PRODUCT, row)
    temp_name = sheet[product_cell_name].value
    if temp_name == 'Марм60ЛТСетМарМоркСв':
        product_name = 'Мармелад овощной'
    elif temp_name == 'Дол100ОвЛТСетМорСвек':
        product_name = 'Дольки овощные'
    elif temp_name == 'Дол100ФрЛТСетЯбГрМан':
        product_name = 'Дольки фруктовые'
    elif temp_name == 'Паст50ЛТСетМалинСлив':
        product_name = 'Пастила С/М'
    elif temp_name == 'Паст50ЛТСетКлубнАбр':
        product_name = 'Пастила К/А'
    elif temp_name == 'Паст50ЛТСетЯблГруши':
        product_name = 'Пастила Я/Г'
    elif temp_name == 'Марм60ЛТСетМарАбрМал':
        product_name = 'Мармелад А/М/С'
    elif temp_name == 'Марм60ЛТСетМарКлубн':
        product_name = 'Мармелад клубника'
    elif temp_name == 'Марм60ЛТСетМарЯблБаз':
        product_name = 'Мармелад яблоко-базилик'
    elif temp_name == 'Паст50ЛТСетнКлубн':
        product_name = 'Пастила клубника'
    elif temp_name == 'Паст50ЛТСетнМалин':
        product_name = 'Пастила малина'
    elif temp_name == 'Печ40ЛТСетнКлубн':
        product_name = 'Печенье клубника'
    elif temp_name == 'Печ40ЛТСетнАбрик':
        product_name = 'Печенье абрикос'
    elif temp_name == 'Печ40ЛТСетнМалин':
        product_name = 'Печенье малина'
    elif temp_name == 'Печ40ЛТСетнСлив':
        product_name = 'Печенье слива'
    elif temp_name == 'Марм60ЛТСетнМарМалин':
        product_name = 'Мармелад малина'


    else:
        product_name = sheet[product_cell_name].value




    leftover = "{}{}".format(COLUMN_LEFTOVER, row)
    input = "{}{}".format(COLUMN_INPUT, row)
    date = "{}{}".format(COLUMN_DATE, row)
    if product_name not in shop:
        product = shop[product_name] = {}
        product['leftover'] = sheet[leftover].value
        product['date'] = sheet[date].value
        product['input'] = sheet[input].value


# new data
for row in range(2, new_data.max_row + 1):
    shop_name_cell = "{}{}".format(COLUMN_SHOP, row)
    shop_name = new_data[shop_name_cell].value
    if shop_name is None:
        break
    new_row_num += 1
    if shop_name not in new_shops:
        new_shops[shop_name] = {}
    shop = new_shops[shop_name]
    add_item(new_data, shop, row)

# old_data
for row in range(2, old_data.max_row + 1):
    shop_name_cell = "{}{}".format(COLUMN_SHOP, row)
    shop_name = old_data[shop_name_cell].value
    if shop_name is None:
        break
    old_row_num += 1
    if shop_name not in old_shops:
        old_shops[shop_name] = {}
    shop = old_shops[shop_name]
    add_item(old_data, shop, row)

# result
result = {}
for shop_name in new_shops:
    result_shop = result[shop_name] = {}
    new_shop = new_shops[shop_name]
    if shop_name in old_shops:
        old_shop = old_shops[shop_name]
    else:
        old_shop = new_shop
    for product_name in new_shop:
        new_product = new_shop[product_name]
        if new_product['date'] is None:
            new_product['date'] = "Пусто"
        if new_product['leftover'] is None:
            new_product['leftover'] = "Пусто"
        if new_product['input'] is None:
            new_product['input'] = "Пусто"

        if new_product['input'] == "Пусто" or new_product['date'] == "Пусто":
            result_product = result_shop[product_name] = {'sells': 0, 'leftover': new_product['leftover'],
                                                          'date': "Пусто", 'input': "Пусто"}
        else:
            result_product = result_shop[product_name] = {'sells': 0, 'leftover': new_product['leftover'],
                                                      'date': new_product['date'], 'input': new_product['input']}
        if product_name not in old_shop:

            if new_product['date'] == "Пусто":
                timediff = new_day - new_day
            else:
                timediff = new_day - new_product['date']

            if timediff.total_seconds() < days_past_in_sec:
                if (new_product['input'] == "Пусто"):
                    result_product['sells'] = 0
                else:
                    result_product['sells'] = new_product['input']
            if not(new_product['leftover'] == "Пусто"):
                result_product['sells'] -= new_product['leftover']

        else:
            old_product = old_shop[product_name]
            if old_product['leftover'] is None:
                old_product['leftover'] = "Пусто"

            if new_product['date'] == "Пусто":
                timediff = new_day - new_day
            else:
                timediff = new_day - new_product['date']

            if timediff.total_seconds() < days_past_in_sec:
                if new_product['input'] == "Пусто":
                    result_product['sells'] = 0
                else:
                    result_product['sells'] = new_product['input']

            if old_product['leftover'] == "Пусто":
                result_product['sells'] += (0 - new_product['leftover'])
            else:
                result_product['sells'] += (old_product['leftover'] - new_product['leftover'])


result_name = 'Звіт за ' + str(new_day.day) + '.' + str(new_day.month)  +'.' + str(new_day.year) + '.xlsx'
workbook = xlsxwriter.Workbook(result_name)
worksheet = workbook.add_worksheet('Data')
workbook.formats[0].set_font_size(14)
bold = workbook.add_format({'size' : 14, 'bold': True, 'text_wrap' : True, 'align' : 'center'})
bold.set_align('vcenter')

colored = workbook.add_format({'size' : 14, 'bold' : True, 'bg_color' : '#d9d9d9', 'border' : True})
colored_n_right = workbook.add_format({'size' : 14, 'bold' : True, 'bg_color' : '#d9d9d9', 'border' : True, 'align' : 'right'})
right = workbook.add_format({'size' : 14, 'align' : 'right'})

row_num = 0

worksheet.set_column(RESULT_SHOP_NAME, RESULT_SHOP_NAME, 22)
worksheet.set_column(RESULT_PROD_NAME, RESULT_PROD_NAME, 25)
worksheet.set_column(RESULT_LEFTOVER, RESULT_LEFTOVER, 9)
worksheet.set_column(RESULT_DATE, RESULT_DATE, 12)
worksheet.set_column(RESULT_INPUT, RESULT_INPUT, 12)
worksheet.set_column(RESULT_SELLS, RESULT_SELLS, 9)
worksheet.set_row(row_num, None, bold)

worksheet.write(row_num, RESULT_SHOP_NAME,  'Магазин', bold)
worksheet.write(row_num, RESULT_PROD_NAME, 'Товар', bold)
worksheet.write(row_num, RESULT_LEFTOVER, 'Остаток', bold)
worksheet.write(row_num, RESULT_DATE, 'Дата последней поставки', bold)
worksheet.write(row_num, RESULT_INPUT, 'Кол-во последнего прихода', bold)
worksheet.write(row_num, RESULT_SELLS, 'Продажи', bold)


row_num += 1
for shop_name in result:
    shop = result[shop_name]
    for prod_name in shop:
        prod = shop[prod_name]
        worksheet.write(row_num, RESULT_SHOP_NAME, shop_name)
        if prod['date'] == "Пусто":
            date = "Пусто"
        else:
            date = str(prod['date'].day) + '.' + str(prod['date'].month) + '.' + str(prod['date'].year)
        if prod['sells'] != 0 or prod['leftover'] != 0:
            worksheet.write(row_num, RESULT_PROD_NAME, prod_name, colored)
            worksheet.write(row_num, RESULT_LEFTOVER, prod['leftover'], colored)
            worksheet.write(row_num, RESULT_DATE, date, colored_n_right)
            worksheet.write(row_num, RESULT_INPUT, prod['input'], colored)
            worksheet.write(row_num, RESULT_SELLS, prod['sells'], colored)
        else:
            worksheet.write(row_num, RESULT_PROD_NAME, prod_name)
            worksheet.write(row_num, RESULT_LEFTOVER, prod['leftover'])
            worksheet.write(row_num, RESULT_DATE, date, right)
            worksheet.write(row_num, RESULT_INPUT, prod['input'])
            worksheet.write(row_num, RESULT_SELLS, prod['sells'])

        row_num += 1
    row_num += 1

worksheet.autofilter(0, RESULT_SHOP_NAME, row_num + 10, RESULT_SELLS)
worksheet.freeze_panes(1, 0)
workbook.close()

